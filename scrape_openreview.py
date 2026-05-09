"""
ICLR 2026 accepted papers scraper.

Fetches metadata, resolves author institutions via OpenReview profiles,
and writes a spreadsheet (xlsx + csv) with: title, authors, institutions,
abstract, decision (Oral/Spotlight/Poster).

Usage:
    python iclr2026_scraper.py --out ./iclr2026
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Iterable

import openreview
from openreview.api import OpenReviewClient
from tqdm import tqdm


VENUE_ID = "ICLR.cc/2026/Conference"
SUBMISSION_INVITATION = f"{VENUE_ID}/-/Submission"
ACCEPTED_MARKERS = ("Poster", "Spotlight", "Oral")


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("iclr2026")


def get_client() -> OpenReviewClient:
    username = os.environ.get("OPENREVIEW_USERNAME")
    password = os.environ.get("OPENREVIEW_PASSWORD")
    return OpenReviewClient(
        baseurl="https://api2.openreview.net",
        username=username,
        password=password,
    )


def fetch_all_submissions(client: OpenReviewClient) -> list[Any]:
    log.info("Fetching submissions via invitation %s", SUBMISSION_INVITATION)
    notes = client.get_all_notes(invitation=SUBMISSION_INVITATION)
    if not notes:
        log.info("Invitation query empty. Falling back to venueid content query.")
        notes = client.get_all_notes(content={"venueid": VENUE_ID})
    log.info("Total notes returned: %d", len(notes))
    return notes


def _content_value(note: Any, key: str, default: Any = None) -> Any:
    content = getattr(note, "content", None) or {}
    field = content.get(key)
    if field is None:
        return default
    if isinstance(field, dict) and "value" in field:
        return field["value"]
    return field


def decision_bucket(venue: str | None) -> str | None:
    if not venue:
        return None
    for marker in ("Oral", "Spotlight", "Poster"):
        if marker in venue:
            return marker
    return None


def normalize_note(note: Any) -> dict[str, Any]:
    venue = _content_value(note, "venue")
    return {
        "id": note.id,
        "forum": getattr(note, "forum", None),
        "title": _content_value(note, "title"),
        "abstract": _content_value(note, "abstract"),
        "authors": _content_value(note, "authors", []) or [],
        "authorids": _content_value(note, "authorids", []) or [],
        "keywords": _content_value(note, "keywords", []) or [],
        "primary_area": _content_value(note, "primary_area"),
        "venue": venue,
        "venueid": _content_value(note, "venueid"),
        "decision": decision_bucket(venue),
    }


def write_jsonl(records: Iterable[dict[str, Any]], path: Path) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    n = 0
    with path.open("w", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
            n += 1
    return n


def fetch_profile_institution(
    client: OpenReviewClient, profile_id: str, cache: dict[str, str]
) -> str:
    """
    Resolve one authorid to an institution string. Tries profile lookup first;
    falls back to email-domain-as-institution. Caches per id.
    """
    if profile_id in cache:
        return cache[profile_id]

    inst = ""
    try:
        if profile_id.startswith("~"):
            profile = client.get_profile(profile_id)
        elif "@" in profile_id:
            profile = client.get_profile(email=profile_id)
        else:
            profile = client.get_profile(profile_id)

        content = getattr(profile, "content", {}) or {}
        history = content.get("history") or []
        if history:
            # Pick the most recent (end is None/empty means current) then fall back to the first
            current = [h for h in history if not h.get("end")]
            chosen = current[0] if current else history[0]
            inst_block = chosen.get("institution") or {}
            inst = inst_block.get("name") or inst_block.get("domain") or ""
        if not inst and "@" in profile_id:
            inst = profile_id.split("@", 1)[1]
    except Exception:
        if "@" in profile_id:
            inst = profile_id.split("@", 1)[1]
        else:
            inst = ""

    cache[profile_id] = inst
    return inst


def enrich_with_institutions(
    client: OpenReviewClient,
    records: list[dict[str, Any]],
    workers: int = 8,
) -> None:
    unique_ids: list[str] = sorted({aid for r in records for aid in r["authorids"] if aid})
    log.info("Resolving institutions for %d unique authorids", len(unique_ids))

    cache: dict[str, str] = {}
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(fetch_profile_institution, client, aid, cache): aid for aid in unique_ids}
        with tqdm(total=len(futures), desc="profiles") as pbar:
            for fut in as_completed(futures):
                try:
                    fut.result()
                except Exception:
                    pass
                pbar.update(1)

    for r in records:
        r["institutions"] = [cache.get(aid, "") for aid in r["authorids"]]


def write_spreadsheet(records: list[dict[str, Any]], out_dir: Path) -> None:
    """Write both CSV and XLSX with the columns the user asked for."""
    import csv

    rows = []
    for r in records:
        authors = r.get("authors") or []
        institutions = r.get("institutions") or []
        rows.append(
            {
                "Decision": r.get("decision") or "",
                "Title": r.get("title") or "",
                "Authors": "; ".join(authors),
                "Institutions": "; ".join([i for i in institutions if i]),
                "Abstract": r.get("abstract") or "",
                "Primary Area": r.get("primary_area") or "",
                "Keywords": "; ".join(r.get("keywords") or []),
                "OpenReview URL": f"https://openreview.net/forum?id={r['forum']}" if r.get("forum") else "",
            }
        )

    rows.sort(key=lambda x: ({"Oral": 0, "Spotlight": 1, "Poster": 2}.get(x["Decision"], 9), x["Title"].lower()))

    columns = ["Decision", "Title", "Authors", "Institutions", "Abstract", "Primary Area", "Keywords", "OpenReview URL"]

    csv_path = out_dir / "iclr2026_accepted.csv"
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=columns)
        w.writeheader()
        w.writerows(rows)
    log.info("Wrote %s (%d rows)", csv_path, len(rows))

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "ICLR 2026 Accepted"
        ws.append(columns)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="305496")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        for r in rows:
            ws.append([r[c] for c in columns])
        widths = {"Decision": 11, "Title": 60, "Authors": 40, "Institutions": 40, "Abstract": 80, "Primary Area": 22, "Keywords": 30, "OpenReview URL": 45}
        for idx, col in enumerate(columns, start=1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = widths[col]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        xlsx_path = out_dir / "iclr2026_accepted.xlsx"
        wb.save(xlsx_path)
        log.info("Wrote %s", xlsx_path)
    except Exception as e:
        log.warning("XLSX export failed (%s). CSV is still written.", e)


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--out", type=Path, default=Path("./iclr2026"))
    p.add_argument("--workers", type=int, default=8)
    p.add_argument("--skip-institutions", action="store_true", help="faster; leaves institutions blank")
    args = p.parse_args()

    args.out.mkdir(parents=True, exist_ok=True)

    client = get_client()
    notes = fetch_all_submissions(client)
    if not notes:
        log.error("No notes returned. Check venue id or auth.")
        return 1

    all_records = [normalize_note(n) for n in notes]
    write_jsonl(all_records, args.out / "submissions_all.jsonl")

    accepted = [r for r in all_records if r["decision"]]
    log.info("Accepted: %d (of %d)", len(accepted), len(all_records))

    buckets: dict[str, int] = {}
    for r in all_records:
        key = r["venue"] or "UNKNOWN/Withdrawn"
        buckets[key] = buckets.get(key, 0) + 1
    for k, v in sorted(buckets.items(), key=lambda kv: -kv[1])[:15]:
        log.info("  %6d  %s", v, k)

    if not args.skip_institutions:
        enrich_with_institutions(client, accepted, workers=args.workers)
    else:
        for r in accepted:
            r["institutions"] = []

    write_jsonl(accepted, args.out / "submissions_accepted.jsonl")
    write_spreadsheet(accepted, args.out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
